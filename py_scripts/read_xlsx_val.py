import sys
import csv
from openpyxl import load_workbook
import re

if (float(sys.version_info[0]*10000)+float(sys.version_info[1])*100+float(sys.version_info[2])) < 20710:
  raise Exception("Blaaast! Python 2.7.10 is required")

def convert_to_csv(ws_active):
  dimensions = ws_active.calculate_dimension()
  if (dimensions == None):
     ws_active.max_row = ws_active.max_column = None
  
  dimensions = ws_active.calculate_dimension()
  print('INFO  : Processing Sheet {0} with dimensions {1}'.format(ws_active.title, dimensions)) 

  if (re.search('A1:A1', dimensions)):
    print('INFO  : Skipping empty sheet ... {0}'.format(ws_active.title))
    return

  with open (ws_active.title +'.csv' , 'wb') as csv_file:
    csv_f = csv.writer(csv_file, delimiter=';')
    for row in ws_active.rows:
      csv_f.writerow([cell.value for cell in row])
  
      for cell in row:
        #if (cell.value != None):
        #print(cell.value)
        pass
  return(ws_active.title)

if (len(sys.argv) == 3):
  xlsx_file = sys.argv[1]
  xlsx_sheet = sys.argv[2]
else:
  print('ERORR : read_xlsx_val.py <workbook.xlsx> <worksheetname|-all|-active>')
  sys.exit(1)

# Get the filname from command line -f
# open the worksheet as read only and to read cached data from formula
# evaluations
# Not sure if this is will be good always.
try:
  wb = load_workbook(filename=xlsx_file, read_only=True, data_only=True)
except IOError:
  print('ERROR : Cannot open workbook....{0}'.format(xlsx_file))

with open ('job_list', 'w') as job_file:
  # Get the sheet name from command line -s or process all (or active...?)
  # Process active could be beneficial for design-render cycles 
  # If sheet name provided, process that sheet
  # else process active sheet (this is the sheet in focus on MSXL)
  # unless -all is specified.
  if(re.search('^-all', xlsx_sheet)):
    sheets = wb.sheetnames
    print('INFO  :Found the following sheets in {0}'.format(xlsx_file))
    print ('     {0}\n'.format(sheets))
    for ws_active in wb:
      if not (re.search(r'_nt$', ws_active.title)):
        result = convert_to_csv(ws_active)
        if (result != None):
          job_file.write(result + ' ')
  
  elif(re.search('^-active', xlsx_sheet)):
    ws_active = wb.active
    ws = ws_active
    print('INFO  :Converting active sheet to csv : {0}\n'.format(ws_active.title))
    result = convert_to_csv(ws_active)
    if (result != None):
      job_file.write(result+' ')
  else:
    if(xlsx_sheet in wb.sheetnames):
      ws_active = wb.get_sheet_by_name(xlsx_sheet)
      result = convert_to_csv(ws_active)
      if (result != None):
        job_file.write(result+' ')
    else:
      print('ERROR : Worksheet not found in workbook \n{0}\n'.format(wb.sheetnames))
      sys.exit(1)


